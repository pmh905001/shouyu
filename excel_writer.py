import logging
import math
import os
import time
from typing import Union

import openpyxl
from PIL.Image import Image as PILImage
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.utils.cell import coordinate_from_string
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from config import ConfigManager
from excel_context import ExcelContext
from msg_box import MessageBox, MessageType
from process import ProcessManager


class ExcelWriter:
    IMAGE_PATH = 'temp.png'

    def __init__(self, excel_path=ConfigManager.excel_path()):
        self._excel_path = excel_path
        self._worksheet_name = time.strftime('%Y-%m-%d')
        self._workbook: Workbook = self._load_workbook()
        self._changed = False

    def _load_workbook(self) -> Workbook:
        if not os.path.exists(self._excel_path):
            workbook = openpyxl.Workbook()
        else:
            workbook = openpyxl.load_workbook(self._excel_path)
        return workbook

    @property
    def _active_worksheet(self) -> Worksheet:
        if self._worksheet_name not in self._workbook.sheetnames:
            worksheet: Worksheet = self._workbook.create_sheet(self._worksheet_name)
            worksheet['A1'] = 'plan'
            self._changed = True
            self._workbook.active = worksheet
        else:
            worksheet: Worksheet = self._workbook.get_sheet_by_name(self._worksheet_name)
            self._workbook.active = worksheet
        return worksheet

    def current_anchor(self, worksheet: Union[str, Worksheet]) -> str:
        if isinstance(worksheet, str):
            worksheet: Worksheet = self._workbook.get_sheet_by_name(worksheet)
        elif isinstance(worksheet, Worksheet):
            worksheet: Worksheet = worksheet
        else:
            raise RuntimeError(f'Invalid worksheet')

        max_image = self._find_max_image(worksheet)
        if max_image:
            if worksheet.max_row < max_image.anchor._from.row + math.ceil(max_image.height / 18) + 1:
                return max_image.anchor, max_image
            else:
                anchor = f'{get_column_letter(self._find_max_column_index(worksheet))}{worksheet.max_row}'
                return anchor, self._active_worksheet[anchor].value
        else:
            if worksheet.max_row == 1 and worksheet.max_column == 1 and worksheet[1][0].value is None:
                anchor = 'A1'
                return anchor, self._active_worksheet[anchor].value
            else:
                anchor = f'{get_column_letter(self._find_max_column_index(worksheet))}{worksheet.max_row}'
                return anchor, self._active_worksheet[anchor].value

    def _next_anchor(self, worksheet: Union[str, Worksheet], column_offset: int = 0, row_offset: int = 0) -> str:
        if isinstance(worksheet, str):
            worksheet: Worksheet = self._workbook.get_sheet_by_name(worksheet)
        elif isinstance(worksheet, Worksheet):
            worksheet: Worksheet = worksheet
        else:
            raise RuntimeError(f'Invalid worksheet')

        max_image = self._find_max_image(worksheet)
        if max_image:
            if worksheet.max_row < max_image.anchor._from.row + math.ceil(max_image.height / 18) + 1:
                return f'{get_column_letter(max_image.anchor._from.col + 1 + column_offset)}{max_image.anchor._from.row + math.ceil(max_image.height / 18) + 1 + row_offset}'
            else:
                return f'{get_column_letter(self._find_max_column_index(worksheet) + column_offset)}{worksheet.max_row + 1 + row_offset}'
        else:
            if worksheet.max_row == 1 and worksheet.max_column == 1 and worksheet[1][0].value is None:
                return f'A{1 + row_offset}'
            else:
                return f'{get_column_letter(self._find_max_column_index(worksheet) + column_offset)}{worksheet.max_row + 1 + row_offset}'

    @staticmethod
    def _find_max_image(sheet: Worksheet) -> Image:
        if not sheet._images:
            return None
        if len(sheet._images) == 1:
            return sheet._images[0]
        max = sheet._images[0]
        for img in sheet._images[1:]:
            if img.anchor._from.row + math.ceil(img.height / 18) > max.anchor._from.row + math.ceil(max.height / 18):
                max = img
        return max

    @staticmethod
    def _find_max_column_index(sheet: Worksheet) -> int:
        if sheet.max_column == 1:
            return 1
        row = sheet[sheet.max_row]
        for i in range(sheet.max_column, 0, -1):
            if row[i - 1].value:
                return i

        return 1

    def _save_image(self, img: PILImage, anchor: str):
        img.save(self.IMAGE_PATH)
        self._active_worksheet.add_image(Image(self.IMAGE_PATH), anchor)
        logging.info('saved image!')

    def _save_text(self, txt: str, anchor: str):
        self._active_worksheet[anchor] = txt
        logging.info(f'saved text: {txt}!')

    def save(self, data: PILImage or str, ignore_permission_error: bool = False, open_excel_again=True):
        msg = None
        duration = ConfigManager.shortcut('save_clipboard_popup_duration', '1', lambda x: int(x))
        image_path = os.path.abspath(self.IMAGE_PATH) if isinstance(data, PILImage) else None
        try:
            if data:
                anchor = self._next_anchor(
                    self._active_worksheet,
                    ExcelContext.get_steps_and_reset(),
                    ExcelContext.get_row_steps_and_reset()
                )
                if isinstance(data, PILImage):
                    self._save_image(data, anchor)
                    msg = f'{anchor}: Image'
                    self._workbook.save(self._excel_path)
                    MessageBox.pop_up_message('Success', msg, MessageType.SUCCESS, duration, image_path)
                else:
                    self._save_text(data, anchor)
                    msg = f'{anchor}: {str(data)}'
                    self._workbook.save(self._excel_path)
                    MessageBox.pop_up_message('Success', msg, MessageType.SUCCESS, duration)
                # Tray._icon.notify(str(data), 'Success')
            else:
                logging.info(f'Nothing to save!')
        except PermissionError as permission_ex:
            if not ignore_permission_error:
                logging.info(f'try to terminate process for {self._excel_path}')
                ProcessManager.terminate_by_path(self._excel_path)
                try:
                    self._workbook.save(self._excel_path)
                    MessageBox.pop_up_message('Success', msg, MessageType.SUCCESS, duration, image_path)
                except Exception as ex:
                    logging.exception(f'failed to save "{data}"')
                    MessageBox.pop_up_message('Failed', str(ex), MessageType.ERROR, duration, image_path)
                if open_excel_again:
                    ProcessManager.open(self._excel_path)
            else:
                logging.exception(f'failed to save "{data}"')
                MessageBox.pop_up_message('Failed', str(permission_ex), MessageType.ERROR, duration, image_path)
        except Exception as ex:
            logging.exception(f'failed to save "{data}"')
            MessageBox.pop_up_message('Failed', str(ex), MessageType.ERROR, duration, image_path)

    def move_column(self, step=0):
        anchor_or_image = self.current_anchor(self._active_worksheet)
        ExcelContext.steps += step
        old = coordinate_from_string(self._next_anchor(self._active_worksheet))[0]
        column_index = column_index_from_string(old)
        if column_index + ExcelContext.steps < 1:
            ExcelContext.steps = 1 - column_index
        if column_index + ExcelContext.steps > 16384:
            ExcelContext.steps = 16384 - column_index

        # if ExcelContext.steps:
        logging.info(f'move {ExcelContext.steps} steps')
        MessageBox.pop_up_message(
            self._generate_move_message(column_index, ExcelContext.steps),
            self._generate_status_message(anchor_or_image),
            MessageType.SUCCESS,
            duration=ConfigManager.shortcut('show_status_popup_duration', '2', lambda x: int(x)),
            image_path=os.path.abspath(self.IMAGE_PATH) if isinstance(anchor_or_image[1], Image) else None
        )

        if self._changed:
            self._workbook.save(self._excel_path)

    def insert_row_sperator(self, step=0):
        ExcelContext.row_steps += step

    def move_to_home(self):
        old = coordinate_from_string(self._next_anchor(self._active_worksheet))[0]
        column_index = column_index_from_string(old)
        ExcelContext.steps = 1 - column_index

        if ExcelContext.steps:
            logging.info(f'move {ExcelContext.steps} steps')
        MessageBox.pop_up_message(
            'Move',
            self._generate_move_message(column_index, ExcelContext.steps),
            MessageType.SUCCESS,
            duration=ConfigManager.shortcut('show_status_popup_duration', '2', lambda x: int(x))
        )

        if self._changed:
            self._workbook.save(self._excel_path)

    @staticmethod
    def _generate_move_message(column_index: int, steps: int):
        from_column = get_column_letter(column_index)
        to_column = get_column_letter(column_index + ExcelContext.steps)
        if steps > 0:
            return f'Move {from_column} -> {to_column}'
        elif steps == 0:
            return from_column
        else:
            return f'Move {to_column} <- {from_column}'

    @staticmethod
    def _generate_status_message(anchor_or_image):
        if isinstance(anchor_or_image[1], Image):
            anchor = anchor_or_image[0]._from
            return f'{get_column_letter(anchor.col + 1) + str(anchor.row)}: Image'
        else:
            return f'{anchor_or_image[0]}:{anchor_or_image[1]}'
