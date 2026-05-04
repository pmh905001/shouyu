"""Plan / Task domain logic backed by the daily Excel tab.

Layout of a daily worksheet:

    Row 1   A: "plan"
    Row 2   B: <task 1 text>     <- font color encodes status
    Row 3   B: <task 2 text>
    Row 4   B: <task 3 text>
    ...     B: more tasks (variable length)
    Row N+1 (empty separator)
    Row N+2 A: <in_progress task title>
    Row N+3 B: detail
    ...     B: detail / C: image / etc.
    (empty separator)
    Row M   A: <next in_progress task title>
    ...

Status is encoded by font color (and a couple of decorations):

    pending      gray
    in_progress  red, bold
    done         green, strike-through
"""
from __future__ import annotations

import logging
import math
from dataclasses import dataclass, field
from enum import Enum
from typing import List, Optional

from openpyxl.styles import Font
from openpyxl.worksheet.worksheet import Worksheet


# Pixels per Excel row used to translate image height -> spanned rows. This
# matches the constant `_next_anchor` already uses in `excel.py`.
_PIXELS_PER_ROW = 18


class TaskStatus(str, Enum):
    PENDING = "pending"
    IN_PROGRESS = "in_progress"
    DONE = "done"


class TaskPriority(str, Enum):
    """Eisenhower-ish priority. NONE means "not yet decided"."""

    NONE = ""
    P1 = "P1"  # must-do today (will hurt if skipped)
    P2 = "P2"  # should-do (high value, recoverable if skipped)
    P3 = "P3"  # nice-to-have

    @classmethod
    def from_value(cls, raw) -> "TaskPriority":
        if raw is None:
            return cls.NONE
        text = str(raw).strip().upper()
        for member in (cls.P1, cls.P2, cls.P3):
            if text == member.value:
                return member
        return cls.NONE


# openpyxl exposes colors as ARGB hex strings ("00RRGGBB") in most cases.
PENDING_COLOR = "FF808080"
IN_PROGRESS_COLOR = "FFC00000"
DONE_COLOR = "FF107C10"

PLAN_HEADER_CELL = "A1"
PLAN_HEADER_TEXT = "plan"
PLAN_TASK_COLUMN = "B"
PLAN_DURATION_COLUMN = "C"
PLAN_PRIORITY_COLUMN = "D"
PLAN_FIRST_TASK_ROW = 2

ACTIVE_TASK_COLUMN = "A"
ACTIVE_DETAIL_COLUMN = "B"
SEPARATOR_ROWS_AFTER_PLAN = 1
SEPARATOR_ROWS_BETWEEN_TASKS = 1

DEFAULT_PLAN_TASKS = ["task 1", "task 2", "task 3"]
POMODORO_LOG_PREFIX = "🍅"


def _font_for(status: TaskStatus) -> Font:
    if status == TaskStatus.IN_PROGRESS:
        return Font(color=IN_PROGRESS_COLOR, bold=True)
    if status == TaskStatus.DONE:
        return Font(color=DONE_COLOR, strike=True)
    return Font(color=PENDING_COLOR)


def _normalize_color(value) -> Optional[str]:
    if value is None:
        return None
    if isinstance(value, str):
        return value.upper()
    rgb = getattr(value, "rgb", None)
    if isinstance(rgb, str):
        return rgb.upper()
    return None


def _status_from_cell(cell) -> TaskStatus:
    font = cell.font
    color = _normalize_color(getattr(font, "color", None))
    if color == IN_PROGRESS_COLOR:
        return TaskStatus.IN_PROGRESS
    if color == DONE_COLOR:
        return TaskStatus.DONE
    return TaskStatus.PENDING


@dataclass
class PlanTask:
    text: str
    status: TaskStatus = TaskStatus.PENDING
    row: int = 0  # 0 means "not yet placed in Excel"
    duration_minutes: int = 0  # 0 means not specified
    priority: TaskPriority = TaskPriority.NONE

    def display_text(self) -> str:
        return self.text or ""


@dataclass
class ActiveEntry:
    """A row in the active area whose A column holds an in_progress / pending task title."""

    row: int
    text: str
    status: TaskStatus = TaskStatus.PENDING
    plan_row: int = 0
    detail_rows: List[int] = field(default_factory=list)


class PlanService:
    """Read / write the plan section and the active task area on a single daily worksheet."""

    def __init__(self, worksheet: Worksheet):
        self.ws = worksheet

    # ---------- plan area ----------

    def ensure_header(self) -> None:
        cell = self.ws[PLAN_HEADER_CELL]
        if cell.value != PLAN_HEADER_TEXT:
            cell.value = PLAN_HEADER_TEXT

    def seed_default_plan(self) -> None:
        """Write the default placeholders if the plan area is completely empty."""
        self.ensure_header()
        if self.read_plan_tasks():
            return
        for i, text in enumerate(DEFAULT_PLAN_TASKS):
            row = PLAN_FIRST_TASK_ROW + i
            cell = self.ws[f"{PLAN_TASK_COLUMN}{row}"]
            cell.value = text
            cell.font = _font_for(TaskStatus.PENDING)

    def read_plan_tasks(self) -> List[PlanTask]:
        tasks: List[PlanTask] = []
        row = PLAN_FIRST_TASK_ROW
        while True:
            cell = self.ws[f"{PLAN_TASK_COLUMN}{row}"]
            value = cell.value
            if value is None or str(value).strip() == "":
                break
            duration_value = self.ws[f"{PLAN_DURATION_COLUMN}{row}"].value
            duration = 0
            if isinstance(duration_value, (int, float)):
                duration = int(duration_value)
            elif isinstance(duration_value, str):
                stripped = duration_value.strip().rstrip("m")
                if stripped.isdigit():
                    duration = int(stripped)
            priority = TaskPriority.from_value(
                self.ws[f"{PLAN_PRIORITY_COLUMN}{row}"].value
            )
            tasks.append(
                PlanTask(
                    text=str(value),
                    status=_status_from_cell(cell),
                    row=row,
                    duration_minutes=duration,
                    priority=priority,
                )
            )
            row += 1
        return tasks

    def write_plan_tasks(self, tasks: List[PlanTask]) -> None:
        """Persist tasks into the plan area, clearing any orphan rows below."""
        self.ensure_header()
        existing = self.read_plan_tasks()
        max_row_to_clear = max(
            [t.row for t in existing] + [PLAN_FIRST_TASK_ROW + len(tasks) - 1, PLAN_FIRST_TASK_ROW]
        )

        for i, task in enumerate(tasks):
            row = PLAN_FIRST_TASK_ROW + i
            text_cell = self.ws[f"{PLAN_TASK_COLUMN}{row}"]
            text_cell.value = task.text
            text_cell.font = _font_for(task.status)
            duration_cell = self.ws[f"{PLAN_DURATION_COLUMN}{row}"]
            duration_cell.value = task.duration_minutes if task.duration_minutes > 0 else None
            priority_cell = self.ws[f"{PLAN_PRIORITY_COLUMN}{row}"]
            priority_cell.value = task.priority.value if task.priority != TaskPriority.NONE else None
            task.row = row

        for row in range(PLAN_FIRST_TASK_ROW + len(tasks), max_row_to_clear + 1):
            for column in (PLAN_TASK_COLUMN, PLAN_DURATION_COLUMN, PLAN_PRIORITY_COLUMN):
                cell = self.ws[f"{column}{row}"]
                cell.value = None
                cell.font = Font()

    def plan_end_row(self) -> int:
        """Return the last row that holds plan content; PLAN_FIRST_TASK_ROW - 1 if empty."""
        tasks = self.read_plan_tasks()
        if not tasks:
            return PLAN_FIRST_TASK_ROW - 1
        return tasks[-1].row

    # ---------- active area ----------

    def active_area_start_row(self) -> int:
        return self.plan_end_row() + 1 + SEPARATOR_ROWS_AFTER_PLAN

    def _row_has_any_content(self, row: int) -> bool:
        max_col = max(self.ws.max_column, 3)
        for col_idx in range(1, max_col + 1):
            if self.ws.cell(row=row, column=col_idx).value is not None:
                return True
        return False

    def last_used_row_in_active_area(self) -> int:
        """Return the largest row index in the active area that's occupied —
        either by cell text OR by an image's visual footprint.

        Without the image-aware part, appending a pomodoro detail right after
        a screenshot would land *on top of* the image (because images don't
        write text into the cells they overlap, so a pure cell-text scan
        misses them). See excel.py:_next_anchor for the same calculation.

        If the active area is empty, returns active_area_start_row() - 1.
        """
        start = self.active_area_start_row()
        last = start - 1
        end = max(self.ws.max_row, start)
        for row in range(start, end + 1):
            if self._row_has_any_content(row):
                last = row

        # Account for any image whose top anchor is inside the active area —
        # its bottom row must be treated as "used" so the next append goes
        # below the image rather than on top of it.
        for img in getattr(self.ws, '_images', None) or []:
            try:
                # openpyxl row anchors are 0-indexed; convert to Excel-style 1-indexed.
                top = img.anchor._from.row + 1
                height_rows = max(1, math.ceil((img.height or 0) / _PIXELS_PER_ROW))
                bottom = top + height_rows - 1
            except Exception:
                logging.exception("failed to compute image footprint")
                continue
            if top >= start and bottom > last:
                last = bottom
        return last

    def list_active_entries(self) -> List[ActiveEntry]:
        """All rows in the active area whose A column holds task titles, in order."""
        entries: List[ActiveEntry] = []
        start = self.active_area_start_row()
        end = max(self.ws.max_row, start)
        current: Optional[ActiveEntry] = None
        for row in range(start, end + 1):
            a_cell = self.ws.cell(row=row, column=1)
            if a_cell.value is not None and str(a_cell.value).strip():
                current = ActiveEntry(
                    row=row,
                    text=str(a_cell.value),
                    status=_status_from_cell(a_cell),
                )
                entries.append(current)
            elif current is not None:
                if self._row_has_any_content(row):
                    current.detail_rows.append(row)
        return entries

    def current_in_progress_entry(self) -> Optional[ActiveEntry]:
        for entry in self.list_active_entries():
            if entry.status == TaskStatus.IN_PROGRESS:
                return entry
        return None

    # ---------- task switching ----------

    def switch_in_progress(self, task: PlanTask) -> int:
        """Demote any current in_progress, then append a new in_progress row in the active area.

        Returns the row number of the newly-written A column cell.
        """
        previous = self.current_in_progress_entry()
        if previous is not None:
            self._set_status(previous.row, TaskStatus.PENDING, column=ACTIVE_TASK_COLUMN)

        last_used = self.last_used_row_in_active_area()
        start = self.active_area_start_row()
        if last_used < start:
            target_row = start
        else:
            target_row = last_used + 1 + SEPARATOR_ROWS_BETWEEN_TASKS

        cell = self.ws.cell(row=target_row, column=1)
        cell.value = task.text
        cell.font = _font_for(TaskStatus.IN_PROGRESS)

        if task.row:
            self._set_status(task.row, TaskStatus.IN_PROGRESS, column=PLAN_TASK_COLUMN)

        logging.info(f"switch_in_progress -> {ACTIVE_TASK_COLUMN}{target_row}: {task.text}")
        return target_row

    def mark_plan_done(self, plan_row: int) -> None:
        if plan_row:
            self._set_status(plan_row, TaskStatus.DONE, column=PLAN_TASK_COLUMN)

    def mark_active_done(self, active_row: int) -> None:
        self._set_status(active_row, TaskStatus.DONE, column=ACTIVE_TASK_COLUMN)

    def _set_status(self, row: int, status: TaskStatus, column: str) -> None:
        cell = self.ws[f"{column}{row}"]
        cell.font = _font_for(status)

    def count_pomodoros_logged(self) -> int:
        """Count active-area cells that look like a logged pomodoro line."""
        start = self.active_area_start_row()
        end = max(self.ws.max_row, start)
        count = 0
        for row in range(start, end + 1):
            for col_idx in range(1, max(self.ws.max_column, 3) + 1):
                value = self.ws.cell(row=row, column=col_idx).value
                if isinstance(value, str) and value.strip().startswith(POMODORO_LOG_PREFIX):
                    count += 1
                    break
        return count
