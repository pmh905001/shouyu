import glob
import logging
import os
import shutil
from io import BytesIO
from typing import List, Optional, Union
from zipfile import BadZipFile

import math
import openpyxl
import time
from PIL.Image import Image as PILImage
from openpyxl.drawing.image import Image
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.utils.cell import coordinate_from_string, coordinate_to_tuple
from openpyxl.utils.units import pixels_to_EMU
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from shouyu.config import Config
from shouyu.decorator.servicehandler import service_handler
from shouyu.service.context import ExcelContext
from shouyu.service.plan import (
    BACKLOG_HEADER_CELL,
    BACKLOG_HEADER_TEXT,
    BACKLOG_SHEET_LIFE,
    BACKLOG_SHEET_WORK,
    OTHER_DETAIL_COLUMN,
    BacklogService,
    PlanService,
    TaskCategory,
)
from shouyu.util.process import ProcessManager


class KbExcel:
    IMAGE_PATH = '../../temp.png'
    POPUP_MSG_LENGTH = 100

    def __init__(self, excel_path=None):
        self._excel_path = excel_path or Config.excel_path()
        self._worksheet_name = time.strftime('%Y-%m-%d')
        # Track whether we had to recover from backup so the caller can warn UX-side.
        self._recovered_from_backup: Optional[str] = None
        self._workbook: Workbook = self._load_workbook()
        self._changed = False
        self._pop_up_msgs = None

    def _load_workbook(self) -> Workbook:
        if not os.path.exists(self._excel_path):
            return openpyxl.Workbook()

        try:
            workbook = openpyxl.load_workbook(self._excel_path)
        except (KeyError, BadZipFile, ValueError) as e:
            # Common corruption signature is `KeyError: '[Content_Types].xml'`
            # caused by a previous in-place save being interrupted. Try to
            # recover automatically from the most recent good backup so the
            # user is never locked out of their own data.
            logging.error(
                f"main Excel file looks corrupt ({e}); attempting backup recovery"
            )
            workbook = self._recover_from_backup()
            if workbook is None:
                logging.warning(
                    "no usable backup found; starting from a fresh empty workbook"
                )
                workbook = openpyxl.Workbook()

        # See `_materialize_images` docstring for why this is critical.
        self._materialize_images(workbook)
        return workbook

    def _recover_from_backup(self) -> Optional[Workbook]:
        """Walk backups newest-first, return the first one that opens cleanly."""
        for backup in self.list_backups(self._excel_path):
            try:
                wb = openpyxl.load_workbook(backup)
                self._recovered_from_backup = backup
                logging.warning(f"auto-recovered workbook from backup: {backup}")
                return wb
            except Exception:
                logging.warning(f"backup is also unreadable, trying next: {backup}")
                continue
        return None

    @staticmethod
    def _materialize_images(workbook: Workbook) -> None:
        """Copy every image's bytes into an in-memory BytesIO right after load.

        openpyxl stores image refs as `ZipExtFile` objects bound to the
        workbook's underlying zip archive. That archive is closed (or GC'd)
        at unpredictable times. The next save then explodes with
        `ValueError: I/O operation on closed file.` partway through writing
        the new xlsx — leaving the destination half-written and corrupt.

        Materializing once on load decouples image data from the loaded zip
        handle, so subsequent saves are safe.
        """
        for ws in workbook.worksheets:
            images = getattr(ws, "_images", None) or []
            for img in list(images):
                try:
                    ref = getattr(img, "ref", None)
                    if ref is None:
                        continue
                    if isinstance(ref, (str, bytes, BytesIO)):
                        continue
                    if hasattr(ref, "read"):
                        try:
                            ref.seek(0)
                        except Exception:
                            pass
                        data = ref.read()
                        if data:
                            img.ref = BytesIO(data)
                except Exception:
                    logging.exception("failed to materialize image into memory")

    @property
    def recovered_from_backup(self) -> Optional[str]:
        """Path of the backup we auto-recovered from, or None for a normal load."""
        return self._recovered_from_backup

    def _regroup_pinned_sheets(self) -> None:
        """Keep the configured pinned sheets (Config.pinned_sheet_names(),
        e.g. account/work/todo list/backlog-work/backlog-life) grouped
        together at the very end of the tab bar, in that exact order, right
        after today's date tab - instead of drifting further and further
        behind it as new daily tabs pile up. Generalizes what used to be a
        "todo list"-only hardcoded rule."""
        pinned = [name for name in Config.pinned_sheet_names() if name in self._workbook.sheetnames]
        if not pinned:
            return
        if self._workbook.sheetnames[-len(pinned):] == pinned:
            return  # already in the right place, in the right order
        for name in pinned:
            worksheet = self._workbook[name]
            self._workbook.move_sheet(worksheet, offset=len(self._workbook.sheetnames) - 1)
        self._changed = True

    @property
    def _active_worksheet(self) -> Worksheet:
        if self._worksheet_name not in self._workbook.sheetnames:
            worksheet: Worksheet = self._workbook.create_sheet(self._worksheet_name)
            PlanService(worksheet).seed_default_plan()
            self._changed = True
            self._workbook.active = worksheet
        else:
            worksheet: Worksheet = self._workbook.get_sheet_by_name(self._worksheet_name)
            self._workbook.active = worksheet

        # Must run *after* today's tab is ensured above: create_sheet()
        # always appends at the absolute end, so regrouping first would just
        # get immediately undone by today's brand-new tab landing after it.
        self._regroup_pinned_sheets()

        for ws in self._workbook.worksheets:
            expected = (ws == worksheet)
            if ws.views.sheetView[0].tabSelected != expected:
                ws.views.sheetView[0].tabSelected = expected
                self._changed = True

        return worksheet

    def current_anchor(self, worksheet: Union[str, Worksheet]) -> str:
        if isinstance(worksheet, str):
            worksheet: Worksheet = self._workbook.get_sheet_by_name(worksheet)
        elif isinstance(worksheet, Worksheet):
            worksheet: Worksheet = worksheet
        else:
            raise RuntimeError(f'Invalid worksheet')

        max_image = self._find_max_image(worksheet)
        if max_image:
            if worksheet.max_row < max_image.anchor._from.row + math.ceil(max_image.height / 18) + 1:
                return max_image.anchor, max_image
            else:
                anchor = f'{get_column_letter(self._find_max_column_index(worksheet))}{worksheet.max_row}'
                return anchor, self._active_worksheet[anchor].value
        else:
            if worksheet.max_row == 1 and worksheet.max_column == 1 and worksheet[1][0].value is None:
                anchor = 'A1'
                return anchor, self._active_worksheet[anchor].value
            else:
                anchor = f'{get_column_letter(self._find_max_column_index(worksheet))}{worksheet.max_row}'
                return anchor, self._active_worksheet[anchor].value

    def _next_anchor(self, worksheet: Union[str, Worksheet], column_offset: int = 0, row_offset: int = 0) -> str:
        if isinstance(worksheet, str):
            worksheet: Worksheet = self._workbook.get_sheet_by_name(worksheet)
        elif isinstance(worksheet, Worksheet):
            worksheet: Worksheet = worksheet
        else:
            raise RuntimeError(f'Invalid worksheet')

        # Every generic append (clipboard/screenshot paste) lands in the
        # "other" scratch area (see plan.py's OTHER_* constants), never in
        # the plan/active area - otherwise `read_plan_tasks()` (which just
        # scans column B until the first blank cell) misreads it as a task.
        # `next_other_row()` already accounts for existing text AND image
        # footprints *within that area*, so it fully replaces the old
        # "find the sheet's overall last row/column" heuristic this used to
        # use. That old heuristic is exactly what caused pasted text to land
        # in column A instead of B: it picked up whichever column the
        # sheet's last row happened to have content in, and right after the
        # 'other' header is created that's column A (the header itself),
        # not this area's own content column B.
        row = PlanService(worksheet).next_other_row() + row_offset
        column = column_index_from_string(OTHER_DETAIL_COLUMN) + column_offset
        return f'{get_column_letter(column)}{row}'

    @staticmethod
    def _find_max_image(sheet: Worksheet) -> Image:
        if not sheet._images:
            return None
        if len(sheet._images) == 1:
            return sheet._images[0]
        max = sheet._images[0]
        for img in sheet._images[1:]:
            if img.anchor._from.row + math.ceil(img.height / 18) > max.anchor._from.row + math.ceil(max.height / 18):
                max = img
        return max

    @staticmethod
    def _find_max_column_index(sheet: Worksheet) -> int:
        if sheet.max_column == 1:
            return 1
        row = sheet[sheet.max_row]
        for i in range(sheet.max_column, 0, -1):
            if row[i - 1].value:
                return i

        return 1

    def _append_image(self, img: PILImage, anchor: str, image_path: Optional[str] = None):
        # openpyxl's Image keeps a *path* reference and only reads bytes off
        # disk at save() time (not at construction time) - so callers that
        # stage more than one image before a single save() MUST give each
        # image its own path, or every embedded image ends up showing
        # whichever file happened to be written last. `self.IMAGE_PATH` (one
        # fixed shared path) is only safe for the single-image-then-save-
        # immediately case; batch dispatchers should pass a unique path.
        path = image_path or self.IMAGE_PATH
        img.save(path)
        self._active_worksheet[anchor]=f'Image created at {time.strftime("%Y-%m-%d %H:%M:%S")}'
        image_obj = Image(path)
        self._active_worksheet.add_image(image_obj, anchor)
        # `add_image()` just stores the raw anchor string as-is
        # (openpyxl.worksheet.worksheet.Worksheet.add_image: `img.anchor =
        # anchor`) - it's only normalized into a real OneCellAnchor object
        # at save() time, and even then the result isn't written back onto
        # the Image (openpyxl.drawing.spreadsheet_drawing._check_anchor
        # returns a new anchor, it doesn't mutate obj.anchor). So an image
        # inspected again in the SAME in-memory session before any save -
        # e.g. staging several screenshots into one batch, see
        # service/dispatch.py - would crash on `.anchor._from` in
        # _find_max_image/_next_anchor below. Normalize immediately so this
        # image behaves exactly like one freshly loaded from disk.
        row, col = coordinate_to_tuple(anchor.upper())
        normalized_anchor = OneCellAnchor()
        normalized_anchor._from.row = row - 1
        normalized_anchor._from.col = col - 1
        normalized_anchor.ext.width = pixels_to_EMU(image_obj.width)
        normalized_anchor.ext.height = pixels_to_EMU(image_obj.height)
        image_obj.anchor = normalized_anchor
        logging.info('saved image!')

    def _append_text(self, txt: str, anchor: str):
        if ExcelContext.cross_multiple_rows:
            col, row = coordinate_from_string(anchor)
            for line in txt.splitlines():
                self._active_worksheet[f'{col}{row}'] = line
                row += 1

        else:
            self._active_worksheet[anchor] = txt
        logging.info(f'saved text: {txt}!')

    @service_handler
    def append(self, data: PILImage or str or List[PILImage or str]):
        if not data:
            logging.info(f'Nothing to save!')
            return

        target_col = ExcelContext.get_target_column_and_reset()

        if isinstance(data, (list, tuple)):
            for record in data:
                self.append_one_record(record, target_col)
        else:
            self.append_one_record(data, target_col)

    def append_one_record(self, data, target_column=None, image_path=None):
        if not data:
            logging.info(f'not save empty or none data!')
            return

        anchor = self._next_anchor(
            self._active_worksheet,
            ExcelContext.get_column_steps_and_reset(),
            ExcelContext.get_row_steps_and_reset()
        )
        if target_column:
            _, row = coordinate_from_string(anchor)
            anchor = f'{target_column}{row}'

        if isinstance(data, PILImage):
            self._append_image(data, anchor, image_path=image_path)
            msg = f'{anchor}: Image'
        else:
            self._append_text(data, anchor)
            msg = f'{anchor}: {str(data)}'
        self._changed = True
        self._pop_up_msgs = {
            'title': 'Submitting',
            'msg': msg[:self.POPUP_MSG_LENGTH],
            'image_path': os.path.abspath(self.IMAGE_PATH) if isinstance(data, PILImage) else None
        }

    @service_handler
    def move_column(self, step=0):
        anchor_or_image = self.current_anchor(self._active_worksheet)
        ExcelContext.column_steps += step
        old = coordinate_from_string(self._next_anchor(self._active_worksheet))[0]
        column_index = column_index_from_string(old)
        if column_index + ExcelContext.column_steps < 1:
            ExcelContext.column_steps = 1 - column_index
        if column_index + ExcelContext.column_steps > 16384:
            ExcelContext.column_steps = 16384 - column_index

        logging.info(f'move {ExcelContext.column_steps} steps')
        self._pop_up_msgs = {
            'title': self._generate_move_message(column_index, ExcelContext.column_steps),
            'msg': self._generate_status_message(anchor_or_image),
            'image_path': os.path.abspath(self.IMAGE_PATH) if isinstance(anchor_or_image[1], Image) else None
        }

    def _backup_excel(self):
        if not os.path.exists(self._excel_path):
            return

        excel_dir = os.path.dirname(self._excel_path) or '.'
        name, ext = os.path.splitext(os.path.basename(self._excel_path))
        timestamp = time.strftime('%Y%m%d_%H%M%S')
        backup_path = os.path.join(excel_dir, f'{name}_backup_{timestamp}{ext}')
        shutil.copy2(self._excel_path, backup_path)
        logging.info(f'backup created: {backup_path}')

        pattern = os.path.join(excel_dir, f'{name}_backup_*{ext}')
        backups = sorted(glob.glob(pattern), key=os.path.getmtime)
        max_backups = Config.max_backups()
        while len(backups) > max_backups:
            oldest = backups.pop(0)
            os.remove(oldest)
            logging.info(f'removed old backup: {oldest}')

    def _save_changed(self):
        if not self._changed:
            return
        self._backup_excel()
        self._atomic_save()
        # Reset so subsequent calls on the same instance don't redundantly save.
        # This matters now that callers (e.g. habit-dialog retry worker) reuse
        # one KbExcel instance across multiple save attempts.
        self._changed = False
        if self._recovered_from_backup is not None:
            logging.info(
                f"persisted recovered workbook back to {self._excel_path}; "
                f"original recovery source was {self._recovered_from_backup}"
            )
            self._recovered_from_backup = None

    def _atomic_save(self) -> None:
        """Save to a temp file then atomically replace the original.

        If save fails midway (PermissionError, the openpyxl image bug, anything),
        the original file stays untouched. This is the single most important
        safety net — without it a partial save corrupts the canonical Excel.

        On PermissionError (target locked by another program), we DO NOT
        auto-rename the tmp to `.unsaved_*.xlsx`. Callers that want that
        last-resort preservation should call `preserve_unsaved()` themselves
        (typically after their retry budget is exhausted) — otherwise every
        retry attempt would leak a `.unsaved_*` file.
        """
        excel_dir = os.path.dirname(self._excel_path) or '.'
        name, ext = os.path.splitext(os.path.basename(self._excel_path))
        tmp_path = os.path.join(excel_dir, f'.{name}.tmp_{os.getpid()}{ext}')
        try:
            self._workbook.save(tmp_path)
        except Exception:
            self._safe_remove(tmp_path)
            raise
        try:
            os.replace(tmp_path, self._excel_path)
        except Exception:
            self._safe_remove(tmp_path)
            raise

    def preserve_unsaved(self) -> Optional[str]:
        """Last-resort: write the current in-memory workbook to a sibling
        `<name>.unsaved_<ts>.xlsx` so the user never silently loses changes
        even when the canonical file remains locked. Returns the path written
        or None on failure.
        """
        excel_dir = os.path.dirname(self._excel_path) or '.'
        name, ext = os.path.splitext(os.path.basename(self._excel_path))
        kept = os.path.join(
            excel_dir,
            f'{name}.unsaved_{time.strftime("%Y%m%d_%H%M%S")}{ext}',
        )
        try:
            self._workbook.save(kept)
            logging.error(f"unsaved changes preserved to: {kept}")
            return kept
        except Exception:
            logging.exception("failed to preserve unsaved changes")
            return None

    @staticmethod
    def _safe_remove(path: str) -> None:
        try:
            if os.path.exists(path):
                os.remove(path)
        except Exception:
            logging.exception(f"failed to remove temp file: {path}")

    @staticmethod
    def list_backups(excel_path: str) -> List[str]:
        """Return all `*_backup_*.xlsx` files for `excel_path`, newest first."""
        if not excel_path:
            return []
        excel_dir = os.path.dirname(excel_path) or '.'
        name, ext = os.path.splitext(os.path.basename(excel_path))
        pattern = os.path.join(excel_dir, f'{name}_backup_*{ext}')
        try:
            return sorted(glob.glob(pattern), key=os.path.getmtime, reverse=True)
        except Exception:
            logging.exception("failed to list backups")
            return []

    @classmethod
    def restore_from_backup(cls, backup_path: str, excel_path: Optional[str] = None) -> str:
        """Restore the canonical Excel file from `backup_path`.

        Before overwriting, the *current* canonical file (which the user might
        still want — e.g. if they picked the wrong backup) is itself backed up
        with a `pre_restore_<ts>` suffix so this operation is reversible.

        Returns the pre-restore safety-backup path (empty string if none).
        Raises on failure.
        """
        target = excel_path or Config.excel_path()
        if not backup_path or not os.path.isfile(backup_path):
            raise FileNotFoundError(f"backup not found: {backup_path}")

        # Validate the backup actually opens before we touch anything.
        try:
            openpyxl.load_workbook(backup_path).close()
        except Exception as e:
            raise RuntimeError(f"backup file is itself unreadable: {e}") from e

        safety_path = ''
        if os.path.exists(target):
            target_dir = os.path.dirname(target) or '.'
            tname, text = os.path.splitext(os.path.basename(target))
            safety_path = os.path.join(
                target_dir,
                f'{tname}.pre_restore_{time.strftime("%Y%m%d_%H%M%S")}{text}',
            )
            try:
                shutil.copy2(target, safety_path)
                logging.info(f"pre-restore safety copy saved to: {safety_path}")
            except Exception:
                logging.exception("failed to take pre-restore safety copy")
                safety_path = ''

        shutil.copy2(backup_path, target)
        logging.info(f"restored Excel from backup: {backup_path} -> {target}")
        return safety_path

    @staticmethod
    def _generate_move_message(column_index: int, steps: int):
        from_column = get_column_letter(column_index)
        to_column = get_column_letter(column_index + ExcelContext.column_steps)
        mode = '' if ExcelContext.cross_multiple_rows else ' & Content in one cell'
        to_row = '' if ExcelContext.row_steps == 0 else f' & Jump {ExcelContext.row_steps} Rows'
        if steps > 0:
            return f'Move {from_column} -> {to_column}{to_row}{mode}'
        elif steps == 0:
            return f'{from_column}{to_row}{mode}'
        else:
            return f'Move {to_column} <- {from_column}{to_row}{mode}'

    @staticmethod
    def _generate_status_message(anchor_or_image):
        if isinstance(anchor_or_image[1], Image):
            anchor = anchor_or_image[0]._from
            return f'{get_column_letter(anchor.col + 1) + str(anchor.row)}: Image'
        else:
            return f'{anchor_or_image[0]}:{anchor_or_image[1]}'

    @classmethod
    def append_title_to_next_row(cls, title: str):
        instance = cls()
        anchor_or_image = instance.current_anchor(instance._active_worksheet)
        if isinstance(anchor_or_image[1], Image):
            anchor = anchor_or_image[0]._from
            current_row = anchor.row + math.ceil(anchor_or_image[1].height / 18)
        else:
            _, current_row = coordinate_from_string(anchor_or_image[0])
        target_cell = f"A{current_row + 1}"
        instance._active_worksheet[target_cell] = title
        instance._changed = True
        instance._pop_up_msgs = {
            "title": "Plan Updated",
            "msg": f"已写入标题: {title} -> {target_cell}",
            "image_path": None,
        }
        instance._save_changed()
        from shouyu.view.msgbox import MessageBox
        MessageBox.pop_up_message(**instance._pop_up_msgs)

    def plan_service(self) -> PlanService:
        return PlanService(self._active_worksheet)

    def plan_service_for(self, date_str: str):
        """Return a PlanService bound to <date_str>'s worksheet, or None if it doesn't exist."""
        if not date_str or date_str not in self._workbook.sheetnames:
            return None
        return PlanService(self._workbook[date_str])

    def backlog_service(self, category: TaskCategory) -> BacklogService:
        """Return a BacklogService bound to the work/life backlog sheet,
        lazily creating (and seeding the header of) the sheet on first use."""
        name = BACKLOG_SHEET_WORK if category == TaskCategory.WORK else BACKLOG_SHEET_LIFE
        if name in self._workbook.sheetnames:
            worksheet = self._workbook[name]
        else:
            worksheet = self._workbook.create_sheet(name)
            worksheet[BACKLOG_HEADER_CELL] = BACKLOG_HEADER_TEXT
            self._changed = True
        return BacklogService(worksheet, category)

    def stage_reflection(self, text: str, date_str: str = None) -> None:
        """Write the reflection into the in-memory workbook only — does NOT
        flush to disk. Use when you want to batch the reflection write with
        plan changes and have a single retryable save at the end."""
        date_str = date_str or time.strftime('%Y-%m-%d')
        if 'reflections' in self._workbook.sheetnames:
            ws = self._workbook['reflections']
        else:
            ws = self._workbook.create_sheet('reflections')
            ws['A1'] = 'date'
            ws['B1'] = 'reflection'
        target_row = None
        for row in range(2, ws.max_row + 1):
            if ws.cell(row=row, column=1).value == date_str:
                target_row = row
                break
        if target_row is None:
            target_row = max(ws.max_row, 1) + 1
            ws.cell(row=target_row, column=1, value=date_str)
        ws.cell(row=target_row, column=2, value=text or '')
        self._changed = True

    def write_reflection(self, text: str, date_str: str = None) -> None:
        self.stage_reflection(text, date_str)
        self._save_changed()

    def read_reflection(self, date_str: str = None) -> str:
        date_str = date_str or time.strftime('%Y-%m-%d')
        if 'reflections' not in self._workbook.sheetnames:
            return ''
        ws = self._workbook['reflections']
        for row in range(2, ws.max_row + 1):
            if ws.cell(row=row, column=1).value == date_str:
                return str(ws.cell(row=row, column=2).value or '')
        return ''

    def append_detail(self, text: str, column: str = 'B') -> str:
        """Append a single detail value into the active area at the next available row.

        Used by pomodoro and other auto-loggers that should not pop up message boxes.
        Returns the anchor written to.
        """
        if not text:
            return ''
        plan = self.plan_service()
        last_used = plan.last_used_row_in_active_area()
        start = plan.active_area_start_row()
        target_row = max(last_used + 1, start)
        anchor = f'{column}{target_row}'
        self._active_worksheet[anchor] = text
        self._changed = True
        self._save_changed()
        return anchor

    def force_save(self) -> None:
        """Public helper to flush pending changes regardless of decorator pipeline."""
        self._save_changed()

    def mark_changed(self) -> None:
        self._changed = True

    @property
    def workbook(self) -> Workbook:
        return self._workbook

    @property
    def active_worksheet(self) -> Worksheet:
        return self._active_worksheet
